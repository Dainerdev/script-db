import copy as _copy
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _nombre_estilo_tabla_xlsxwriter(nombre_openpyxl):
    """
    openpyxl guarda el estilo de tabla como "TableStyleMedium2";
    xlsxwriter espera "Table Style Medium 2" (con espacios). Inserta un
    espacio antes de cada mayúscula (que no sea la primera letra) y antes
    de cada dígito.
    """
    if not nombre_openpyxl:
        return None
    return re.sub(r"(?<=[a-z])(?=[A-Z])|(?<=[A-Za-z])(?=[0-9])", " ", nombre_openpyxl)

def export_excel(df, file_path):
    """
    Function to export a DataFrame to an Excel file
    """
    
    try:
        # Export the DataFrame to an Excel file
        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active
        
        fecha_col = None
        
        # Search for the "Fecha" column
        # BUGFIX: la comparacion era case-sensitive ("Fecha" exacto) y la
        # columna real se llama "FECHA" (mayusculas) tanto en el archivo
        # original como despues de standarize_column_dates, asi que el
        # formato DD/MM/YYYY nunca se aplicaba (se veia en Excel como
        # 2019-03-29 en vez de 29/03/2019).
        for cell in ws[1]:
            if str(cell.value).strip().upper() == "FECHA":
                fecha_col = cell.column
                break
        
        # Apply date formatting to the "Fecha" column
        if fecha_col is not None:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, fecha_col)
                
                if cell.value is not None:
                    cell.number_format = "DD/MM/YYYY"
            
        wb.save(file_path)
        
        print(f"\nArchivo de resultados guardado en: {file_path}")
    
    except Exception as e:
        print(f"Error exporting Excel file: {e}")


def read_header_style(source_path, source_sheet=None):
    """
    "Hay alguna forma de copiar el estilo original?"

    Lee del archivo original: fuente y alineación del encabezado, alto de
    la fila 1, y el ancho de cada columna INDEXADO POR NOMBRE (no por
    letra), para poder aplicarlo después aunque el archivo de salida tenga
    columnas nuevas o en otro orden. Para columnas que no existían en el
    original (ej. "Nombres", "Reparto_Categoria") se calcula un ancho
    promedio como valor por defecto, en vez de dejarlas con el ancho
    genérico de openpyxl.
    """
    wb = load_workbook(source_path)
    ws = wb[source_sheet] if source_sheet else wb.active
    header_row = list(ws[1])

    font_ref = _copy.copy(header_row[0].font) if header_row else None
    align_ref = _copy.copy(header_row[0].alignment) if header_row else None
    row_height = ws.row_dimensions[1].height

    anchos_por_nombre = {}
    for cell in header_row:
        if cell.value is not None:
            letra = cell.column_letter
            ancho = ws.column_dimensions[letra].width if letra in ws.column_dimensions else None
            anchos_por_nombre[str(cell.value).strip()] = ancho

    validos = [w for w in anchos_por_nombre.values() if w]
    ancho_default = sum(validos) / len(validos) if validos else None

    # "eso incluye los colores?" -> el encabezado azul + filas con banda
    # celeste NO es relleno manual de celda (fill quedó vacío en la lectura
    # de arriba): el original está definido como una Tabla de Excel con
    # nombre (ws.tables), con un estilo con nombre (ej. "TableStyleMedium2")
    # que Excel renderiza en tiempo real. Para replicar el color hay que
    # replicar la TABLA, no una celda.
    table_style = None
    banded_rows = True
    if ws.tables:
        primera_tabla = ws.tables[list(ws.tables.keys())[0]]
        if primera_tabla.tableStyleInfo:
            table_style = primera_tabla.tableStyleInfo.name
            banded_rows = bool(primera_tabla.tableStyleInfo.showRowStripes)

    return {
        "font": font_ref,
        "alignment": align_ref,
        "row_height": row_height,
        "anchos_por_nombre": anchos_por_nombre,
        "ancho_default": ancho_default,
        "table_style": table_style,
        "banded_rows": banded_rows,
    }


def apply_header_style(ws, columns, style_ref):
    """
    Aplica a `ws` el estilo leído con read_header_style(). Las columnas que
    existían en el archivo original (comparando por nombre) recuperan su
    mismo ancho; las columnas nuevas quedan con el ancho promedio de las
    originales, para no verse desproporcionadas. La fuente y alineación del
    encabezado se aplican a todas las columnas por igual, existan o no en
    el original.
    """
    if style_ref.get("row_height"):
        ws.row_dimensions[1].height = style_ref["row_height"]

    for i, col_name in enumerate(columns, start=1):
        letra = get_column_letter(i)
        cell = ws.cell(1, i)
        if style_ref.get("font"):
            cell.font = _copy.copy(style_ref["font"])
        if style_ref.get("alignment"):
            cell.alignment = _copy.copy(style_ref["alignment"])

        ancho = style_ref["anchos_por_nombre"].get(str(col_name).strip(), style_ref.get("ancho_default"))
        if ancho:
            ws.column_dimensions[letra].width = ancho



def export_multi_sheet_excel(sheets, file_path, style_reference_path=None, style_reference_sheet=None):
    """
    Exporta varios DataFrames a un mismo archivo Excel, un sheet por cada
    entrada de `sheets` (dict: nombre_de_hoja -> DataFrame). Aplica el
    formato DD/MM/YYYY a cualquier columna "FECHA" que encuentre, en
    cualquier hoja. Si se pasa `style_reference_path`, además replica en
    cada hoja el estilo de encabezado (fuente, alineación, alto de fila,
    anchos de columna) de ese archivo -ver read_header_style().

    NOTA DE RENDIMIENTO (dos vueltas):
    1. La primera version releia el archivo completo con load_workbook()
       despues de escribirlo y recorria celda por celda con
       ws.cell(row, col) -no alcanzaba a terminar en una hoja de ~185 mil
       filas, asi que el formato quedaba sin aplicar aunque los datos si se
       guardaban.
    2. La segunda version ya no releia el archivo, pero seguia usando el
       engine 'openpyxl' para ESCRIBIR, que en la prueba tardó 79.6s solo
       para la hoja de 184.692 x 27 -sigue sin alcanzar a terminar dentro
       de un tiempo razonable sumando las otras 7 hojas. 'openpyxl'
       construye un objeto por cada celda; para escrituras grandes es
       conocido por ser mucho más lento que la alternativa.
    Esta version usa el engine 'xlsxwriter' (ya está en requirements.txt)
    para escribir, y aplica formato POR COLUMNA COMPLETA
    (worksheet.set_column) en vez de celda por celda -una sola llamada por
    columna en vez de una por cada fila.
    """
    try:
        style_ref = read_header_style(style_reference_path, style_reference_sheet) if style_reference_path else None
        estilo_tabla = _nombre_estilo_tabla_xlsxwriter(style_ref["table_style"]) if style_ref else None

        # date_format/datetime_format van en el CONSTRUCTOR de ExcelWriter:
        # to_excel() ya escribe las celdas de fecha con su propio formato
        # (por defecto "YYYY-MM-DD") en el momento en que las escribe; un
        # set_column() posterior no alcanza a sobreescribir celdas que ya
        # tienen formato propio. Puesto aquí, aplica a CUALQUIER columna de
        # tipo fecha en CUALQUIER hoja (FECHA, Reparto_Fecha, etc.).
        with pd.ExcelWriter(
            file_path, engine="xlsxwriter",
            date_format="dd/mm/yyyy", datetime_format="dd/mm/yyyy",
        ) as writer:
            workbook = writer.book

            for sheet_name, data in sheets.items():
                safe_name = str(sheet_name)[:31]  # limite de Excel para nombres de hoja

                if estilo_tabla:
                    # el encabezado azul + filas con banda celeste del
                    # original NO es relleno manual de celda, es el estilo
                    # con nombre de una Tabla de Excel (ver
                    # read_header_style). Se escribe la data SIN encabezado
                    # propio (header=False, startrow=1) y se deja que
                    # add_table() cree el encabezado -asi hereda el color y
                    # el filtro automático igual que el original.
                    data.to_excel(writer, sheet_name=safe_name, index=False, header=False, startrow=1)
                    ws = writer.sheets[safe_name]
                    ws.add_table(0, 0, len(data), len(data.columns) - 1, {
                        "style": estilo_tabla,
                        "banded_rows": style_ref.get("banded_rows", True),
                        "columns": [{"header": str(c)} for c in data.columns],
                    })
                else:
                    data.to_excel(writer, sheet_name=safe_name, index=False)
                    ws = writer.sheets[safe_name]

                if style_ref and style_ref.get("row_height"):
                    ws.set_row(0, style_ref["row_height"])

                for i, col_name in enumerate(data.columns):
                    col_str = str(col_name).strip()
                    ancho = style_ref["anchos_por_nombre"].get(col_str, style_ref.get("ancho_default")) if style_ref else None
                    if ancho:
                        ws.set_column(i, i, ancho)

        print(f"\nArchivo con {len(sheets)} hoja(s) guardado en: {file_path}")

    except Exception as e:
        print(f"Error exporting multi-sheet Excel file: {e}")