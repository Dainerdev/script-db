import unicodedata
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook


# CONFIGURATION

# Excel base file
ARCHIVE_EXCEL = "archive/probando.xlsx"

# Diagnostic results file
ARCHIVE_DIAGNOSTIC = "archive/diagnostic_results.xlsx"


# READ EXCEL FILES
def read_excel_file(file_path):
    """
    Function to read an Excel file into a DataFrame
    """
    
    try:
        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_path, engine="openpyxl")
        return df
    
    except FileNotFoundError:
        print(f"Error: File not found - {file_path}")
        return None
    
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None


# FUNCTIONS

# Function to standarize column names
def standarize_column_names(serie):
    """
    Function to standarize column names
    """    
    
    # Ensure the series is of string type
    serie = serie.astype(str)  
    
    # Remove leading and trailing whitespace
    serie = serie.str.strip()
    
    # Replace spaces with underscores
    serie = serie.str.replace(r"\s+", " ", regex=True)  # Replace multiple spaces with a single space
    
    # Convert to uppercase
    serie = serie.str.upper()
    
    # Normalize unicode characters to ASCII
    serie = serie.map(remove_accents)
    
    return serie

# Remove accents from a string
def remove_accents(s):
    """
    Function to remove accents from a string
    """
    
    # Normalize the string to NFKD form and encode to ASCII, ignoring errors
    return unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")

# Function to standarize column dates
def standarize_column_dates(serie):
    """
    Function to standarize column dates
    """
    
    # Ensure the series is of string type and remove leading/trailing whitespace 
    serie = serie.astype(str).str.strip()
    
    dates = pd.to_datetime(serie, errors="coerce", dayfirst=True, format="mixed")
    
    # Normalize to remove time component
    return dates.dt.normalize()

# IMPLMENTATION

def main():
    """
    Main function to execute the diagnostic process
    """
    
    # Load the Excel file
    df = read_excel_file(ARCHIVE_EXCEL)
    
    if df is not None:
        # Display general information about the DataFrame
        rows = len(df)
        columns = len(df.columns)

        print("\nINFORMACIÓN GENERAL\n") 

        print(f"Filas: {rows:,}") 
        print(f"Columnas: {columns:,}") 
        print(f"Tamaño: {rows:,} x {columns:,}")
    
    # Standardize columns
    # in the "Nombres" column of the DataFrame
    df["Nombres"] = standarize_column_names(df["Nombres"])
    
    # in the "Fecha" column of the DataFrame
    df["Fecha"] = standarize_column_dates(df["Fecha"])

    # Save the modified DataFrame to a new Excel file
    df.to_excel(ARCHIVE_DIAGNOSTIC, index=False)

    wb = load_workbook(ARCHIVE_DIAGNOSTIC)
    ws = wb.active
    
    # Search for the "Fecha" column
    for cell in ws[1]:
        if cell.value == "Fecha":
            fecha_col = cell.column
            break
    
    # Apply date formatting to the "Fecha" column
    for row in range(2, ws.max_row + 1):
        ws.cell(row, fecha_col).number_format = "DD/MM/YYYY"
        
    wb.save(ARCHIVE_DIAGNOSTIC)
    
    print(f"\nArchivo de resultados guardado en: {ARCHIVE_DIAGNOSTIC}")
    
if __name__ == "__main__":
    main()
